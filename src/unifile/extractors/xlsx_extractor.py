# Copyright (c) 2025 takotime808

from __future__ import annotations

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from unifile.extractors.base import (
    BaseExtractor,
    make_row,
    Row,
)


class ExcelExtractor(BaseExtractor):
    """XLSX/XLS → text extractor with optional cell-level granularity.

    By default the extractor emits **one row per worksheet** whose content is
    the tab-delimited representation of the sheet.  When ``as_cells`` is ``True``
    each individual cell is emitted as its own row with ``unit_type="cell"`` and
    ``metadata`` describing the sheet, row and column indices.

    Inherits :meth:`BaseExtractor.extract` for path validation and
    exception-to-error-row wrapping. The actual spreadsheet reading is
    implemented in :meth:`_extract`.

    Supported extensions
    --------------------
    xlsx, xlsm, xltx, xltm, xls

    Output Row (per sheet)
    ----------------------
    - file_type: normalized from the file suffix (e.g., "xlsx")
    - unit_type: "sheet"
    - unit_id:   worksheet title
    - content:   Tab-delimited lines (one per spreadsheet row)
    - metadata:  {"nrows": int, "ncols": int}
    """

    supported_extensions = ["xlsx", "xlsm", "xltx", "xltm", "xls"]

    def __init__(self, *, as_cells: bool = False):
        """Parameters
        ----------
        as_cells:
            When ``True`` emit one row per cell instead of one row per sheet.
        """
        self.as_cells = as_cells

    def _extract(self, path: Path) -> List[Row]:
        """
        Read a workbook and return standardized rows, one per worksheet.

        Parameters
        ----------
        path
            Path to an Excel workbook. Existence checks are handled by
            :class:`BaseExtractor.extract`.

        Returns
        -------
        list[Row]
            Standardized rows for each worksheet.
        """
        rows: List[Row] = []
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                if self.as_cells:
                    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        for c_idx, val in enumerate(row, start=1):
                            txt = "" if val is None else str(val)
                            rows.append(
                                make_row(
                                    path=path,
                                    file_type=path.suffix.lstrip(".").lower() or "xlsx",
                                    unit_type="cell",
                                    unit_id=f"{ws.title}!{r_idx},{c_idx}",
                                    content=txt,
                                    metadata={
                                        "sheet": ws.title,
                                        "row": r_idx,
                                        "col": c_idx,
                                    },
                                    status="ok",
                                )
                            )
                else:
                    lines: list[str] = []
                    for row in ws.iter_rows(values_only=True):
                        vals = ["" if v is None else str(v) for v in row]
                        lines.append("\t".join(vals))
                    text = "\n".join(lines)
                    rows.append(
                        make_row(
                            path=path,
                            file_type=path.suffix.lstrip(".").lower() or "xlsx",
                            unit_type="sheet",
                            unit_id=ws.title,
                            content=text,
                            metadata={"nrows": ws.max_row, "ncols": ws.max_column},
                            status="ok",
                        )
                    )
        finally:
            wb.close()
        return rows


class CsvExtractor(BaseExtractor):
    """CSV/TSV → text extractor with optional cell rows.

    The default behaviour emits a single row whose ``content`` is the canonical
    CSV serialization.  When ``as_cells`` is ``True`` each cell becomes an
    individual row similar to :class:`ExcelExtractor`.

    Inherits :meth:`BaseExtractor.extract` for path validation and
    exception-to-error-row wrapping. The actual parsing is implemented in
    :meth:`_extract`.

    Supported extensions
    --------------------
    csv, tsv

    Output Row
    ----------
    - file_type: "csv" or "tsv"
    - unit_type: "table"
    - unit_id:   "0"
    - content:   CSV text (comma-separated; TSV is first read with tab sep)
    - metadata:  {"rows": int, "cols": int}
    """

    supported_extensions = ["csv", "tsv"]

    def __init__(self, *, as_cells: bool = False):
        self.as_cells = as_cells

    def _extract(self, path: Path) -> List[Row]:
        """
        Parse a CSV/TSV file into a single standardized row.

        Parameters
        ----------
        path
            Path to a `.csv` or `.tsv` file. Existence checks are handled by
            :class:`BaseExtractor.extract`.

        Returns
        -------
        list[Row]
            A single row with CSV text content and basic table metadata.
        """
        sep = "\t" if path.suffix.lower().lstrip(".") == "tsv" else ","
        df = pd.read_csv(path, sep=sep, dtype=str)
        if self.as_cells:
            out: List[Row] = []
            for r_idx in range(len(df)):
                for c_idx, col in enumerate(df.columns):
                    val = df.iloc[r_idx, c_idx]
                    txt = "" if pd.isna(val) else str(val)
                    out.append(
                        make_row(
                            path=path,
                            file_type=path.suffix.lstrip(".").lower() or "csv",
                            unit_type="cell",
                            unit_id=f"{r_idx},{c_idx}",
                            content=txt,
                            metadata={"row": r_idx, "col": col},
                            status="ok",
                        )
                    )
            return out

        text = df.to_csv(index=False)
        return [
            make_row(
                path=path,
                file_type=path.suffix.lstrip(".").lower() or "csv",
                unit_type="table",
                unit_id="0",
                content=text,
                metadata={"rows": len(df), "cols": len(df.columns)},
                status="ok",
            )
        ]
